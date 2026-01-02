"""
将指定文件夹下的所有Excel题目生成AI解析
支持多线程并发处理（默认最多10个文件同时进行）
command: python excel_generate_analysis.py --folder ./data
"""

import openpyxl
import requests
import time
import os
import argparse
import glob
from concurrent.futures import ThreadPoolExecutor, as_completed

# ================= 配置区域 =================
parser = argparse.ArgumentParser(description="Excel题目解析生成器(多线程版)")
parser.add_argument(
    "--folder",
    type=str,
    default=r"D:\电网\新增题库",
    help="包含Excel文件的文件夹路径",
)
parser.add_argument(
    "--workers",
    type=int,
    default=25,
    help="同时处理的文件数量（默认10）",
)
config = parser.parse_args()

# 请在此处填入你的 Key


# 列名关键词映射
HEADER_KEYWORDS = {
    "question": ["题目名称", "题目", "题干", "内容"],
    "option_a": ["选项A", "A", "选项 A"],
    "option_content1": ["选项内容1", "选项内容 A"],
    "option_b": ["选项B", "B", "选项 B"],
    "option_content2": ["选项内容2", "选项内容 B"],
    "option_c": ["选项C", "C", "选项 C"],
    "option_content3": ["选项内容3", "选项内容 C"],
    "option_d": ["选项D", "D", "选项 D"],
    "option_content4": ["选项内容4", "选项内容 D"],
    "answer": ["正确答案", "答案", "参考答案"],
    "analysis": ["解析", "题目解析"],
}


# ===========================================
def call_doubao_api(prompt):
    """调用豆包 (火山引擎) 获取解析"""
    if not DOUBAO_API_KEY:
        print("错误: 未配置 DOUBAO_API_KEY")
        return None

    # 火山引擎 (Ark) 的标准兼容接口地址
    url = "https://ark.cn-beijing.volces.com/api/v3/chat/completions"

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {DOUBAO_API_KEY}",
    }

    data = {
        # 注意：这里需要填入【推理接入点 ID】，而不是模型名称
        "model": DOUBAO_ENDPOINT_ID,
        "messages": [
            {
                "role": "system",
                "content": "你是一位计算机辅导老师。请针对题目给出解析。回答简洁明了，别说废话。",
            },
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.7,  # 豆包建议稍微降低一点温度以保证稳定性
        "stream": False,
    }

    for attempt in range(3):
        try:
            # 增加超时时间，豆包有时候处理较慢
            res = requests.post(url, headers=headers, json=data, timeout=60)

            if res.status_code == 200:
                # 豆包的返回结构与 OpenAI/DeepSeek 兼容
                return res.json()["choices"][0]["message"]["content"]
            else:
                # 打印错误详情方便调试
                print(f"⚠️ 豆包报错: {res.status_code} - {res.text}")
                time.sleep(1)
        except Exception as e:
            print(f"网络请求异常: {e}")
            time.sleep(1)

    return None


def call_deepseek_api(prompt):
    """调用 DeepSeek 获取解析"""
    if not Deep_seek_API_KEY:
        return None
    url = "https://api.deepseek.com/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {Deep_seek_API_KEY}",
    }
    data = {
        "model": "deepseek-chat",
        "messages": [
            {
                "role": "system",
                "content": "你是一位计算机辅导老师。请针对题目给出解析。回答简洁明了，别说废话。",
            },
            {"role": "user", "content": prompt},
        ],
        "temperature": 1.0,
        "stream": False,
    }

    for attempt in range(3):
        try:
            res = requests.post(url, headers=headers, json=data, timeout=60)
            if res.status_code == 200:
                return res.json()["choices"][0]["message"]["content"]
            else:
                # print(f"⚠️ DeepSeek 报错: {res.status_code}")
                time.sleep(1)
        except Exception:
            time.sleep(1)
    return None


def call_kimi_api(prompt):
    """调用 Kimi 获取解析"""
    if not Kimi_API_KEY:
        return None
    url = "https://api.moonshot.cn/v1/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {Kimi_API_KEY}",
    }
    data = {
        "model": "moonshot-v1-8k",
        "messages": [
            {
                "role": "system",
                "content": "你是一位计算机辅导老师。请针对题目给出解析。回答简洁明了，别说废话。",
            },
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.3,
        "stream": False,
    }

    for attempt in range(3):
        try:
            res = requests.post(url, headers=headers, json=data, timeout=60)
            if res.status_code == 200:
                return res.json()["choices"][0]["message"]["content"]
            else:
                if res.status_code == 429:
                    time.sleep(5)  # Rate limit backoff
        except Exception:
            time.sleep(1)
    return None


def call_tongyi_judge(question_context, deepseek_ans, kimi_ans, original_ans=None):
    """核心裁判逻辑"""
    if not Tongyi_API_KEY:
        return deepseek_ans if deepseek_ans else kimi_ans

    url = "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {Tongyi_API_KEY}",
    }

    judge_content = f"【题目信息】\n{question_context}\n\n"
    options_map = {}

    if deepseek_ans:
        judge_content += f"【待选解析 A (DeepSeek)】\n{deepseek_ans}\n\n"
        options_map["A"] = deepseek_ans
    if kimi_ans:
        judge_content += f"【待选解析 B (Kimi)】\n{kimi_ans}\n\n"
        options_map["B"] = kimi_ans
    if original_ans and len(str(original_ans)) > 5:
        judge_content += f"【待选解析 C (原始记录)】\n{original_ans}\n\n"
        options_map["C"] = original_ans

    if not options_map:
        return None
    if len(options_map) == 1:
        return list(options_map.values())[0]

    judge_content += """
    请作为该领域的资深专家，评估上述不同来源的解析。
    评判标准：1. 准确性 2. 详尽性 3. 易读性。
    请决策：哪个解析质量最高？
    **请只返回最佳解析对应的字母标签（A、B 或 C），不要包含任何标点符号或其他废话。**
    """

    data = {
        "model": "qwen-plus",
        "messages": [
            {"role": "system", "content": "你是一个只输出标签（A/B/C）的评判机器。"},
            {"role": "user", "content": judge_content},
        ],
        "temperature": 0.1,
    }

    for attempt in range(3):
        try:
            res = requests.post(url, headers=headers, json=data, timeout=60)
            if res.status_code == 200:
                result_tag = (
                    res.json()["choices"][0]["message"]["content"].strip().upper()
                )
                target_key = None
                if "A" in result_tag and "A" in options_map:
                    target_key = "A"
                elif "B" in result_tag and "B" in options_map:
                    target_key = "B"
                elif "C" in result_tag and "C" in options_map:
                    target_key = "C"

                return (
                    options_map[target_key]
                    if target_key
                    else (deepseek_ans or kimi_ans)
                )
        except Exception:
            time.sleep(1)

    return deepseek_ans if deepseek_ans else kimi_ans


def find_column_indices(sheet):
    """映射表头列号"""
    mapping = {}
    for col_idx in range(1, sheet.max_column + 1):
        cell_val = sheet.cell(row=1, column=col_idx).value
        if not cell_val:
            continue
        cell_str = str(cell_val).strip()
        for key, keywords in HEADER_KEYWORDS.items():
            if key not in mapping and cell_str in keywords:
                mapping[key] = col_idx
    return mapping


def process_single_excel(file_path):
    """
    处理单个 Excel 文件的核心逻辑
    """
    filename = os.path.basename(file_path)
    print(f"🔄 [开始处理] {filename}")

    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    except Exception as e:
        print(f"❌ [读取失败] {filename}: {e}")
        return

    col_map = find_column_indices(sheet)
    if "question" not in col_map:
        print(f"⚠️ [跳过] {filename} - 未找到‘题目’列")
        return

    # 确保解析列存在
    if "analysis" not in col_map:
        new_col = sheet.max_column + 1
        sheet.cell(row=1, column=new_col).value = "解析"
        col_map["analysis"] = new_col

    rows = list(sheet.iter_rows(min_row=2))
    total_rows = len(rows)
    processed_count = 0

    # 辅助函数：安全获取单元格值
    def get_val(r_idx, key):
        if key in col_map:
            val = sheet.cell(row=r_idx, column=col_map[key]).value
            return str(val).strip() if val else ""
        return ""

    # 遍历行 (去掉了 tqdm，改用简单的进度打印，因为多线程下 tqdm 会乱)
    for i, row in enumerate(rows):
        row_idx = row[0].row

        # 每处理10条打印一次日志，避免刷屏
        if i % 10 == 0 and i > 0:
            print(f"   ⏳ [{filename}] 进度: {i}/{total_rows}")

        q_text = get_val(row_idx, "question")
        if not q_text or q_text.lower() == "nan":
            continue

        original_analysis = sheet.cell(row=row_idx, column=col_map["analysis"]).value

        correct_answer = get_val(row_idx, "answer")
        prompt_text = f"""
        题目：{q_text}
        选项：
        A. {get_val(row_idx, 'option_a')} {get_val(row_idx, 'option_content1')}
        B. {get_val(row_idx, 'option_b')} {get_val(row_idx, 'option_content2')}
        C. {get_val(row_idx, 'option_c')} {get_val(row_idx, 'option_content3')}
        D. {get_val(row_idx, 'option_d')} {get_val(row_idx, 'option_content4')}
        参考答案：{correct_answer}
        
        要求：
        1. 请给出知识点解析,尽量简洁，别说废话。
        """
        prompt_text = prompt_text.strip()

        # 串行调用 API（每个线程内部串行）
        ds_res = call_deepseek_api(prompt_text)
        ki_res = call_kimi_api(prompt_text)
        # doubao_res = call_doubao_api(prompt_text)
        best_analysis = call_tongyi_judge(
            prompt_text, ds_res, ki_res, original_analysis
        )

        if best_analysis:
            sheet.cell(row=row_idx, column=col_map["analysis"]).value = best_analysis
            processed_count += 1

    # 保存文件
    dir_name = os.path.dirname(file_path)
    final_name = os.path.join(dir_name, f"res_{filename}")
    try:
        wb.save(final_name)
        print(
            f"✅ [完成] {filename} -> 已保存至: {final_name} (处理了 {processed_count} 题)"
        )
    except Exception as e:
        print(f"❌ [保存失败] {filename}: {e}")


def main():
    target_folder = config.folder
    max_workers = config.workers

    if not os.path.isdir(target_folder):
        print(f"❌ 文件夹不存在: {target_folder}")
        return

    # 扫描所有 xlsx 文件
    # 排除已经生成的 res_ 开头的文件，防止循环处理
    all_files = glob.glob(os.path.join(target_folder, "*.xlsx"))
    files_to_process = [
        f
        for f in all_files
        if not os.path.basename(f).startswith("res_")
        and not os.path.basename(f).startswith("~$")
    ]

    print(f"📂 扫描目录: {target_folder}")
    print(f"🔢 发现 Excel 文件: {len(files_to_process)} 个")
    print(f"🚀 启动多线程处理 (最大并发: {max_workers})...\n")

    # 使用线程池并发处理
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 提交所有任务
        futures = {
            executor.submit(process_single_excel, f_path): f_path
            for f_path in files_to_process
        }

        # 等待完成（此处可选）
        for future in as_completed(futures):
            f_path = futures[future]
            try:
                future.result()
            except Exception as exc:
                print(f"❌ 文件 {f_path} 处理过程抛出未捕获异常: {exc}")

    print("\n🎉 所有文件处理任务结束！")


if __name__ == "__main__":
    main()
