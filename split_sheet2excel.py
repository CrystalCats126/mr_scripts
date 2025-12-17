import openpyxl
import os
import argparse

# 设置命令行参数
parser = argparse.ArgumentParser(description="批量将文件夹下的Excel文件按Sheet拆分")
parser.add_argument(
    "-d",
    "--dir",
    type=str,
    default="E:\my_script\题目分类2（南网）",
    help="指定要处理的文件夹路径 (默认为当前目录)",
)
config = parser.parse_args()


def split_single_excel(file_path, output_root_folder):
    """
    处理单个 Excel 文件：读取 Sheet，保留格式拆分，按 '文件名-Sheet名' 保存
    """
    # 1. 获取基础文件名 (用于命名新文件)
    file_basename = os.path.basename(file_path)  # 例如: a.xlsx
    file_name_no_ext = os.path.splitext(file_basename)[0]  # 例如: a

    print(f"--> 正在读取文件：{file_basename}")

    try:
        # 2. 第一次加载：仅为了获取 Sheet 名称列表 (read_only 模式速度快)
        wb_readonly = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = wb_readonly.sheetnames
        wb_readonly.close()

        print(f"    检测到 {len(sheet_names)} 个 Sheet: {sheet_names}")

        # 3. 循环处理每个 Sheet
        for target_sheet in sheet_names:
            # 重新加载完整的工作簿 (为了保留格式，必须 data_only=False)
            # 注意：对于大文件，反复加载会比较慢，但这是openpyxl保留样式的唯一方法
            wb = openpyxl.load_workbook(file_path, data_only=False)

            # 遍历工作簿中的所有 Sheet，删除不需要的
            for sheet in wb.sheetnames:
                if sheet != target_sheet:
                    del wb[sheet]

            # 4. 构建新的文件名：原文件名-Sheet名.xlsx
            new_filename = f"{file_name_no_ext}-{target_sheet}.xlsx"
            output_path = os.path.join(output_root_folder, new_filename)

            # 保存
            wb.save(output_path)
            wb.close()
            print(f"    ✅ 已保存: {new_filename}")

    except Exception as e:
        print(f"    ❌ 处理文件 {file_basename} 时发生错误: {e}")


def process_folder(folder_path):
    """
    遍历文件夹并处理所有 Excel 文件
    """
    # 1. 检查输入路径
    if not os.path.exists(folder_path):
        print(f"错误：找不到文件夹 {folder_path}")
        return

    # 2. 创建统一的输出目录
    output_root_folder = os.path.join(folder_path, "所有拆分结果")
    if not os.path.exists(output_root_folder):
        os.makedirs(output_root_folder)
        print(f"已创建输出目录：{output_root_folder}")

    # 3. 遍历文件夹下的所有文件
    files = [
        f
        for f in os.listdir(folder_path)
        if f.endswith(".xlsx") and not f.startswith("~$")
    ]

    if not files:
        print("该文件夹下没有找到 .xlsx 文件。")
        return

    print(f"共发现 {len(files)} 个 Excel 文件，开始处理...\n" + "=" * 30)

    for file_name in files:
        full_file_path = os.path.join(folder_path, file_name)
        split_single_excel(full_file_path, output_root_folder)
        print("-" * 30)

    print(f"\n🎉 所有任务完成！文件已保存在: {output_root_folder}")


if __name__ == "__main__":
    # 获取参数中的路径，如果没有指定，则使用当前代码所在的目录
    target_dir = config.dir

    # 开始处理
    process_folder(target_dir)
