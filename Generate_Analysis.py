"""
将指定的文件里面的题目生成AI解析
command: python Generate_Analysis.py --file 2021年国家电网考试真题（计算机类）.xlsx
"""

import openpyxl
from openpyxl.utils import get_column_letter
import requests
import time
import os
from tqdm import tqdm
import argparse

# ================= 配置区域 =================
parser = argparse.ArgumentParser(description="Excel题目解析生成器")
parser.add_argument(
    "--file",
    type=str,
    default="2014年国家电网考试真题（计算机类）.xlsx",
    help="Excel文件路径",
)
config = parser.parse_args()

# 请在此处填入你的 Key


TARGET_FILE = config.file

# 列名关键词映射
HEADER_KEYWORDS = {
    "question": ["题目名称", "题目", "题干", "内容"],
    "option_a": ["选项A", "A", "选项 A"],
    "option_content1": ["选项内容1", "选项内容 A"],
    "option_b": ["选项B", "B", "选项 B"],
    "option_content2": ["选项内容2", "选项内容 B"],
    "option_c": ["选项C", "C", "选项 C"],
    "option_content3": ["选项内容3", "选项内容 C"],
    "option_d": ["选项D", "D", "选项 D"],
    "option_content4": ["选项内容4", "选项内容 D"],  # 修正了之前的拼写错误
    "answer": ["正确答案", "答案", "参考答案"],
    "analysis": ["解析", "题目解析"],
}
# ===========================================


def call_deepseek_api(prompt):
    """调用 DeepSeek 获取解析"""
    if not Deep_seek_API_KEY:
        return None
    url = "https://api.deepseek.com/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {Deep_seek_API_KEY}",
    }
    data = {
        "model": "deepseek-chat",
        "messages": [
            {
                "role": "system",
                "content": "你是一位计算机辅导老师。请针对题目给出解析。回答简洁明了，别说废话。",
            },
            {"role": "user", "content": prompt},
        ],
        "temperature": 1.0,
        "stream": False,
    }

    for attempt in range(3):
        try:
            res = requests.post(url, headers=headers, json=data, timeout=60)
            if res.status_code == 200:
                return res.json()["choices"][0]["message"]["content"]
            else:
                print(f"⚠️ DeepSeek 报错: {res.status_code}")
        except Exception as e:
            print(f"⏳ DeepSeek 连接异常: {e}")
        time.sleep(1)
    return None


def call_kimi_api(prompt):
    """调用 Kimi 获取解析"""
    if not Kimi_API_KEY:
        return None
    url = "https://api.moonshot.cn/v1/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {Kimi_API_KEY}",
    }
    data = {
        "model": "moonshot-v1-8k",
        "messages": [
            {
                "role": "system",
                "content": "你是一位计算机辅导老师。请针对题目给出解析。回答简洁明了，别说废话。",
            },
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.3,
        "stream": False,
    }

    for attempt in range(3):
        try:
            res = requests.post(url, headers=headers, json=data, timeout=60)
            if res.status_code == 200:
                return res.json()["choices"][0]["message"]["content"]
            else:
                print(f"⚠️ Kimi 报错: {res.status_code}")
                if res.status_code == 429:
                    time.sleep(3)
        except Exception as e:
            print(f"⏳ Kimi 连接异常: {e}")
        time.sleep(1)
    return None


def call_tongyi_judge(question_context, deepseek_ans, kimi_ans, original_ans=None):
    """
    核心裁判逻辑：
    输入：题目上下文、DeepSeek答案、Kimi答案、原始答案(可选)
    输出：最好的那个答案的文本
    """
    if not Tongyi_API_KEY:
        # 如果没有配置裁判 Key，默认返回 DeepSeek，如果没有则返回 Kimi
        return deepseek_ans if deepseek_ans else kimi_ans

    url = "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {Tongyi_API_KEY}",
    }

    # 构造裁判 Prompt
    judge_content = f"【题目信息】\n{question_context}\n\n"

    options_map = {}  # 用于存储标签和内容的映射

    if deepseek_ans:
        judge_content += f"【待选解析 A (DeepSeek)】\n{deepseek_ans}\n\n"
        options_map["A"] = deepseek_ans

    if kimi_ans:
        judge_content += f"【待选解析 B (Kimi)】\n{kimi_ans}\n\n"
        options_map["B"] = kimi_ans

    if original_ans and len(str(original_ans)) > 5:  # 原始解析太短通常没意义
        judge_content += f"【待选解析 C (原始记录)】\n{original_ans}\n\n"
        options_map["C"] = original_ans

    # 如果没有足够的选项进行比较，直接返回有的那个
    if not options_map:
        return None
    if len(options_map) == 1:
        return list(options_map.values())[0]

    judge_content += """
    请作为该领域的资深专家，评估上述不同来源的解析。
    评判标准：
    1. 准确性：必须符合题目原本的正确答案。
    2. 详尽性：解析是否清晰、逻辑是否闭环。
    3. 易读性：排版整洁。

    请决策：哪个解析质量最高？
    **请只返回最佳解析对应的字母标签（A、B 或 C），不要包含任何标点符号或其他废话。**
    """

    data = {
        "model": "qwen-plus",  # 使用 Plus 或 Max 版本以获得更好的判断力
        "messages": [
            {"role": "system", "content": "你是一个只输出标签（A/B/C）的评判机器。"},
            {"role": "user", "content": judge_content},
        ],
        "temperature": 0.1,
    }

    for attempt in range(3):
        try:
            res = requests.post(url, headers=headers, json=data, timeout=60)
            if res.status_code == 200:
                result_tag = (
                    res.json()["choices"][0]["message"]["content"].strip().upper()
                )
                print(f"   ⚖️ 裁判选择: {result_tag}", end="")

                # 清洗结果，防止模型输出 "选A" 这种包含中文的情况
                target_key = None
                if "A" in result_tag and "A" in options_map:
                    target_key = "A"
                elif "B" in result_tag and "B" in options_map:
                    target_key = "B"
                elif "C" in result_tag and "C" in options_map:
                    target_key = "C"

                if target_key:
                    return options_map[target_key]
                else:
                    print(f" -> 格式异常({result_tag})，默认选 DeepSeek/Kimi")
                    return deepseek_ans if deepseek_ans else kimi_ans
            else:
                print(f"⚠️ 通义裁判报错: {res.status_code}")
        except Exception as e:
            print(f"❌ 裁判连接失败: {e}")
        time.sleep(1)

    # 裁判失败兜底：优先 DeepSeek
    return deepseek_ans if deepseek_ans else kimi_ans


def find_column_indices(sheet):
    """映射表头列号"""
    mapping = {}
    for col_idx in range(1, sheet.max_column + 1):
        cell_val = sheet.cell(row=1, column=col_idx).value
        if not cell_val:
            continue
        cell_str = str(cell_val).strip()
        for key, keywords in HEADER_KEYWORDS.items():
            if key not in mapping and cell_str in keywords:
                mapping[key] = col_idx
    return mapping


def main():
    print(f"📂 加载文件: {TARGET_FILE}")
    try:
        wb = openpyxl.load_workbook(TARGET_FILE)
        sheet = wb.active
    except Exception as e:
        print(f"❌ 无法打开: {e}")
        return

    col_map = find_column_indices(sheet)
    if "question" not in col_map:
        print("❌ 未找到‘题目’列")
        return

    # 确保解析列存在
    if "analysis" not in col_map:
        new_col = sheet.max_column + 1
        sheet.cell(row=1, column=new_col).value = "解析"
        col_map["analysis"] = new_col
        print(f"🆕 新建解析列: 第 {new_col} 列")

    rows = list(sheet.iter_rows(min_row=2))

    # 记录修改状态
    processed_count = 0

    print("🚀 开始多模型竞技场处理...")

    # 使用 tqdm 显示进度
    for row in tqdm(rows):
        row_idx = row[0].row

        # 获取各列内容的辅助函数
        def get_val(key):
            if key in col_map:
                val = sheet.cell(row=row_idx, column=col_map[key]).value
                return str(val).strip() if val else ""
            return ""

        q_text = get_val("question")
        if not q_text or q_text.lower() == "nan":
            continue

        # 1. 构造题目 Prompt
        # 注意：这里我们让 DeepSeek 和 Kimi 既看正确答案（如果有的话），也看选项
        correct_answer = get_val("answer")
        prompt_text = f"""
        题目：{q_text}
        选项：
        A. {get_val('option_a')} {get_val('option_content1')}
        B. {get_val('option_b')} {get_val('option_content2')}
        C. {get_val('option_c')} {get_val('option_content3')}
        D. {get_val('option_d')} {get_val('option_content4')}
        
        {f'参考答案：{correct_answer}' if correct_answer else ''}
        
        要求：
        1. 请给出知识点解析,尽量简洁，别说废话。
        """
        prompt_text = prompt_text.strip()
        # 2. 获取现有解析（如果有）
        original_analysis = sheet.cell(row=row_idx, column=col_map["analysis"]).value
        # 如果已经有很长的解析，你可以选择跳过，或者强制重跑（这里设为强制重跑）
        # if original_analysis and len(str(original_analysis)) > 50: continue

        # 3. 并行调用（此处为简化写成了串行，但因为有 tqdm 监控进度也无妨）
        # 获取 DeepSeek 答案
        ds_res = call_deepseek_api(prompt_text)

        # 获取 Kimi 答案
        ki_res = call_kimi_api(prompt_text)

        # 4. 召唤裁判 (通义千问)
        # 将题目、DeepSeek结果、Kimi结果、原始结果 一起发给裁判
        best_analysis = call_tongyi_judge(
            prompt_text, ds_res, ki_res, original_analysis
        )

        # 5. 写入 Excel
        if best_analysis:
            sheet.cell(row=row_idx, column=col_map["analysis"]).value = best_analysis
            processed_count += 1

    # 最终保存
    final_name = f"res_{TARGET_FILE}"
    try:
        wb.save(final_name)
        print(f"\n✅ 全部完成！结果已保存至: {final_name}")
    except PermissionError:
        print("\n❌ 保存失败：请关闭 Excel 文件后重试。")


if __name__ == "__main__":
    main()
