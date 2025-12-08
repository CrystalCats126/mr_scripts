"""
自动在题库里面搜索题目并且更新解析

使用方法：
python website_auto_fill.py --file "res_通信-第一章信息新技术(1).xlsx" --url https://cwgedu.cn/powerTutoring-ui/#/addAnswerSubjectQuestion/index/297/2
"""

import asyncio
from playwright.async_api import async_playwright
from openpyxl import load_workbook
import argparse

# ================= 配置区域 =================
parser = argparse.ArgumentParser(description="自动上传解析")
parser.add_argument(
    "--file",
    type=str,
    default="res_通信-第四章数据通信网(1).xlsx",
    help="Excel文件路径",
)
parser.add_argument(
    "--url",
    type=str,
    default="https://cwgedu.cn/powerTutoring-ui/#/addAnswerSubjectQuestion/index/294/2",
    help="网站url",
)
config = parser.parse_args()


# ================= 配置区域 =================
# 1. 这里填你【另存为】后的 xlsx 文件名
DATA_FILE = config.file

# 2. 你的网站地址
TARGET_URL = config.url

# 3. 列名配置 (必须和Excel第一行表头完全一致)
COL_QUESTION_NAME = "题目名称"
COL_ANALYSIS_NAME = "解析"
# ===========================================


async def run():
    print(f"正在使用 openpyxl 读取文件: {DATA_FILE}...")

    try:
        # 加载 Excel 文件 (data_only=True 确保读取的是值而不是公式)
        wb = load_workbook(filename=DATA_FILE, data_only=True)
        sheet = wb.active  # 获取第一个工作表
    except Exception as e:
        print(f"❌ 读取文件失败: {e}")
        print(
            "请确认：\n1. 文件名写对了吗？\n2. 文件是 .xlsx 格式吗？(openpyxl 不支持 csv)"
        )
        return

    # === 获取表头，找到“题目名称”和“解析”在第几列 ===
    # openpyxl 的索引从 1 开始
    headers = []
    # 读取第一行作为表头
    for cell in sheet[1]:
        headers.append(str(cell.value).strip() if cell.value else "")

    try:
        # 找到列的索引 (Python列表从0开始，但后面取值时要注意)
        q_index = headers.index(COL_QUESTION_NAME)
        a_index = headers.index(COL_ANALYSIS_NAME)
        print(
            f"✅ 成功找到列：'{COL_QUESTION_NAME}' 在第 {q_index+1} 列，'{COL_ANALYSIS_NAME}' 在第 {a_index+1} 列。"
        )
    except ValueError as e:
        print(f"❌ 错误：在表头中没找到指定的列名。")
        print(f"Excel里的表头是: {headers}")
        return

    async with async_playwright() as p:
        print("正在启动浏览器...")
        # 有头模式 + 慢动作
        browser = await p.chromium.launch(headless=False, slow_mo=800)
        context = await browser.new_context()
        page = await context.new_page()
        await page.set_viewport_size({"width": 1400, "height": 900})

        # === 步骤 1: 人工登录 ===
        await page.goto(TARGET_URL)
        print("\n" + "=" * 50)
        print("【请手动操作】")
        print("请登录并跳转到【题目列表】页面。")
        input(">>> 准备好后，按【回车键】开始自动化...")
        print("=" * 50 + "\n")

        # === 步骤 2: 循环处理 ===
        # min_row=2 表示从第2行开始读（跳过表头）
        # values_only=True 直接获取单元格的值
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        total_rows = len(rows)

        for i, row in enumerate(rows):
            # 获取当前行的题目和解析
            # 注意：row 是一个元组，索引对应之前的 headers 索引
            question_text = (
                str(row[q_index]).strip() if row[q_index] is not None else ""
            )
            analysis_text = (
                str(row[a_index]).strip() if row[a_index] is not None else ""
            )

            # 跳过无效数据
            if not question_text or question_text == "None":
                continue
            if analysis_text == "None":
                analysis_text = ""

            print(f"[{i+1}/{total_rows}] 正在操作: {question_text[:10]}...")

            try:
                # ==========================================
                # --- A. 搜索 (终极稳定版) ---
                # ==========================================

                # 1. 定位：找到包含“题目名称”文字的那个区域，再找里面的input
                print("刷新页面...")
                await page.reload()
                await page.wait_for_timeout(3000)
                # 刷新后重新获取元素
                search_item = page.locator(".el-form-item").filter(has_text="题目名称")
                search_input = search_item.locator("input")

                # 输入新题目
                await search_input.fill(question_text)
                await search_input.press("Enter")

                # 4. 等待加载
                # 搜索后，页面通常会有短暂的加载动画，建议等待一下
                # 如果能找到加载遮罩最好： await page.locator('.el-loading-mask').wait_for(state='hidden')
                await page.wait_for_timeout(1500)

                # --- B. 点击修改 ---
                target_row = page.locator("tr").filter(has_text=question_text)
                edit_btn = target_row.locator('button:has-text("修改")')

                if await edit_btn.count() > 0:
                    await edit_btn.first.scroll_into_view_if_needed()
                    await edit_btn.first.click()
                else:
                    print(f"   -> [跳过] 没找到对应题目: {question_text[:10]}")
                    continue

                # --- C. 填写解析 ---
                analysis_box = page.locator('textarea[placeholder="请输入解析"]')
                await analysis_box.wait_for(state="visible", timeout=5000)

                await analysis_box.clear()
                await analysis_box.fill(analysis_text)

                # --- D. 保存 ---
                # --- D. 保存 (最终修复版) ---
                print("   -> 正在尝试保存...")

                # 1. 锁定弹窗底部区域 (el-dialog__footer)
                # 这一步是为了确保我们点的是弹窗里的按钮，而不是页面背后的
                dialog_footer = page.locator(".el-dialog__footer").filter(
                    has=page.locator(":visible")
                )

                # 2. 定位按钮
                # 方案 A: 优先找蓝色的主按钮 (el-button--primary) 且包含 "确" 字的
                # 注意：这里匹配 "确" 字即可，不管后面有没有空格
                save_btn = dialog_footer.locator("button.el-button--primary").filter(
                    has_text="确"
                )

                # 方案 B: 如果 A 没找到，尝试找文本完全匹配 "确 定" (带空格) 的按钮
                if await save_btn.count() == 0:
                    save_btn = dialog_footer.locator('button:has-text("确 定")')

                # 方案 C: 还是没找到？试试不带空格的 "确定" (防止有些题目不一样)
                if await save_btn.count() == 0:
                    save_btn = dialog_footer.locator('button:has-text("确定")')

                # 3. 执行点击
                if await save_btn.count() > 0:
                    # force=True 强行点击，忽略动画遮挡
                    await save_btn.first.click(force=True)
                else:
                    print("   -> [警告] 实在找不到保存按钮，尝试盲按回车...")
                    await page.keyboard.press("Enter")

                # --- E. 等待完成 ---
                await analysis_box.wait_for(state="hidden", timeout=5000)
                print("   -> [成功]")

            except Exception as e:
                print(f"   -> [错误] {e}")
                await page.reload()
                await page.wait_for_timeout(2000)

    print("\n全部搞定！")


if __name__ == "__main__":
    asyncio.run(run())
